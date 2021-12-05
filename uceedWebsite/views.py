from django.shortcuts import  render, redirect
from .forms import NewUserForm
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.forms import AuthenticationForm,PasswordChangeForm
from django.urls import reverse_lazy
import openpyxl
import json


#The django refrence bookmark folder in chrome has all the tutorials you need to make a website.
def homepage(request):
	return render(request,"homepage.html")

def about(request):
    return render(request,"about.html")

def contact(request):
    return render(request,"contact.html")

def profile(request):
    return render(request,"profile.html")

def flashcard_request(request):
	path = "D:\\anshu\\python\\uceed-website\\venv\\uceedWebsite\\uceedWebsite\\Book1.xlsx"
	wb_obj = openpyxl.load_workbook(path)
	sheet_obj = wb_obj.active
	cell_obj = sheet_obj.cell(row=2, column=1)
	qna = {
		"questions": [],
		'answers': []
	}

	for x in range(2, sheet_obj.max_row + 1):
		qna['questions'].append(sheet_obj.cell(row=x, column=1).value)

	for y in range(2, sheet_obj.max_row + 1):
		qna['answers'].append(sheet_obj.cell(row=y, column=2).value)

	with open('uceedWebsite/templates/data.json', 'w') as outfile:
		json.dump(qna, outfile)

	return render(request,"flashcard.html" )

def register_request(request,backend='django.contrib.auth.backends.ModelBackend'):
	if request.method == "POST":
		form = NewUserForm(request.POST)
		if form.is_valid():
			user = form.save()
			login(request, user, backend='django.contrib.auth.backends.ModelBackend')
			messages.success(request, "Registration successful." )
			return redirect("/")
		messages.error(request, "Unsuccessful registration. Invalid information.")
	form = NewUserForm()
	return render (request=request, template_name="register.html", context={"register_form":form})

def login_request(request, backend='django.contrib.auth.backends.ModelBackend'):
	if request.method == "POST":
		form = AuthenticationForm(request, data=request.POST)
		if form.is_valid():
			username = form.cleaned_data.get('username')
			password = form.cleaned_data.get('password')
			user = authenticate(username=username, password=password)
			if user is not None:
				login(request, user, backend='django.contrib.auth.backends.ModelBackend')
				messages.info(request, f"You are now logged in as {username}.")
				return redirect("/")
			else:
				messages.error(request, "Invalid username or password.")
		else:
			messages.error(request, "Invalid username or password.")
	form = AuthenticationForm()
	return render(request=request, template_name="login.html", context={"login_form": form})

def logout_request(request):
	logout(request)
	messages.info(request, "You have successfully logged out.")
	return redirect("/")

def download(request):
	return render(request, "download.html")

class PasswordsChangeView(PasswordChangeView):
	form_class= PasswordChangeForm
	success_url = reverse_lazy('homepage')